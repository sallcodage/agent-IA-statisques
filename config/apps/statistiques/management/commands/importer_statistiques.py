import csv
import io
import os
from decimal import Decimal

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from config.apps.statistiques.models import StatistiqueRegionale

REQUIRED_FIELDS = [
    "region", "annee", "population", "taux_urbanisation_pct",
    "taux_alphabetisation_pct", "taux_chomage_pct", "taux_pauvrete_pct",
    "acces_internet_pct", "centres_sante", "taux_scolarisation_pct",
    "production_cerealiere_tonnes"
]


def _decode_csv_text(raw_bytes):
    for encoding in ("utf-8-sig", "utf-16", "utf-16-le", "utf-16-be", "cp1252", "latin-1"):
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw_bytes.decode("utf-8-sig", errors="replace")


def _iter_rows_from_csv(path):
    with open(path, "rb") as handle:
        raw_bytes = handle.read()
    text = _decode_csv_text(raw_bytes)
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return []

    normalized_fields = [field.strip() if field else "" for field in reader.fieldnames]
    reader.fieldnames = normalized_fields
    rows = []
    for row in reader:
        cleaned_row = {}
        for field in normalized_fields:
            value = row.get(field)
            if value is None:
                cleaned_row[field] = ""
            elif isinstance(value, str):
                cleaned_row[field] = value.strip()
            else:
                cleaned_row[field] = value
        rows.append(cleaned_row)
    return rows


def _iter_rows_from_excel(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = []
    header = None

    for row in sheet.iter_rows(values_only=True):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        if header is None:
            header = [str(cell).strip() if cell is not None else "" for cell in row]
            continue

        cleaned_row = {}
        for index, field in enumerate(header):
            value = row[index] if index < len(row) else ""
            cleaned_row[field] = "" if value is None else str(value).strip()
        rows.append(cleaned_row)

    workbook.close()
    return rows


class Command(BaseCommand):
    help = "Importe un fichier CSV ou Excel de statistiques régionales"

    def add_arguments(self, parser):
        parser.add_argument("file_path", nargs=1, type=str)

    def handle(self, *args, **options):
        file_path = options["file_path"][0]
        if not os.path.exists(file_path):
            self.stderr.write(self.style.ERROR(f"Fichier introuvable : {file_path}"))
            return

        extension = os.path.splitext(file_path)[1].lower()
        try:
            if extension in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
                rows = _iter_rows_from_excel(file_path)
            else:
                rows = _iter_rows_from_csv(file_path)
        except Exception as exc:
            self.stderr.write(self.style.ERROR(f"Impossible d’importer le fichier : {exc}"))
            return

        if not rows:
            self.stderr.write(self.style.ERROR("Le fichier est vide."))
            return

        fieldnames = list(rows[0].keys())
        missing = [field for field in REQUIRED_FIELDS if field not in fieldnames]
        if missing:
            self.stderr.write(self.style.ERROR(f"Colonnes manquantes : {', '.join(missing)}"))
            return

        created = 0
        updated = 0
        rejected = 0
        for row in rows:
            try:
                cleaned = {
                    "region": row["region"].strip(),
                    "annee": int(row["annee"]),
                    "population": int(row["population"]),
                    "taux_urbanisation_pct": Decimal(str(row["taux_urbanisation_pct"])),
                    "taux_alphabetisation_pct": Decimal(str(row["taux_alphabetisation_pct"])),
                    "taux_chomage_pct": Decimal(str(row["taux_chomage_pct"])),
                    "taux_pauvrete_pct": Decimal(str(row["taux_pauvrete_pct"])),
                    "acces_internet_pct": Decimal(str(row["acces_internet_pct"])),
                    "centres_sante": int(row["centres_sante"]),
                    "taux_scolarisation_pct": Decimal(str(row["taux_scolarisation_pct"])),
                    "production_cerealiere_tonnes": int(row["production_cerealiere_tonnes"]),
                }
            except (ValueError, KeyError, TypeError):
                rejected += 1
                continue

            if not (2020 <= cleaned["annee"] <= 2024):
                rejected += 1
                continue
            for field in [
                "taux_urbanisation_pct", "taux_alphabetisation_pct", "taux_chomage_pct",
                "taux_pauvrete_pct", "acces_internet_pct", "taux_scolarisation_pct"
            ]:
                if not (Decimal("0") <= cleaned[field] <= Decimal("100")):
                    rejected += 1
                    break
            else:
                _, created_or_updated = StatistiqueRegionale.objects.update_or_create(
                    region=cleaned["region"],
                    annee=cleaned["annee"],
                    defaults=cleaned,
                )
                if created_or_updated:
                    created += 1
                else:
                    updated += 1

        self.stdout.write(self.style.SUCCESS(
            f"Import terminé : {created} créées, {updated} mises à jour, {rejected} rejetées"
        ))
